from openpyxl import load_workbook

def add_pipes(destination_sheet, source_sheet):
    row = 2

    while source_sheet[f"B{row}"].value:
        dest_row = 11 + (row - 2)

        destination_sheet[f"B{dest_row}"] = "NEW"
        destination_sheet[f"E{dest_row}"] = "PR"

        original_value = source_sheet[f"B{row}"].value or ""
        destination_sheet[f"D{dest_row}"] = original_value[:32]

        destination_sheet[f"F{dest_row}"] = "BPIPPIPE      " + destination_sheet[f"D{dest_row}"].value
        destination_sheet[f"C{dest_row}"] = source_sheet[f"A{row}"].value
        destination_sheet[f"O{dest_row}"] = source_sheet[f"D{row}"].value
        destination_sheet[f"P{dest_row}"] = source_sheet[f"E{row}"].value

        row += 1

    return 0

def convert_psim_to_asse(psim_file, second_file, output_file):
    workbook_dest = load_workbook(filename=psim_file)
    workbook_src = load_workbook(filename=second_file)

    destination_sheet = workbook_dest['BPIPPIPE']
    source_sheet = workbook_src['Sheet2']

    add_pipes(destination_sheet, source_sheet)

    workbook_dest.save(filename=output_file)
