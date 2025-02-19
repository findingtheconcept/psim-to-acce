from openpyxl import load_workbook
from fileinput import filename

def Add_Pipes(sheet, data):
    row = 2

    while data["B" + str(row)].value:
        sheet["B" + str(11+row-2)] = "NEW"
        sheet["E" + str(11+row-2)] = "PR"
        sheet["D" + str(11+row-2)] = data["B" + str(row)].value[:32]
        sheet["F" + str(11+row-2)] = "BPIPPIPE      " + sheet["D" + str(11+row-2)].value
        sheet["C" + str(11+row-2)] = data["A" + str(row)].value
        sheet["O" + str(11+row-2)] = data["D" + str(row)].value
        sheet["P" + str(11+row-2)] = data["E" + str(row)].value
        row += 1
        print(row)

    return 0

workbook = load_workbook(filename="C:\\Users\\roman\\Desktop\\small_export_pipes.xlsx")
workbook2 = load_workbook(filename="C:\\Users\\roman\\Desktop\\toACCE.xlsx")


Add_Pipes(workbook['BPIPPIPE'], workbook2['Sheet2'])

workbook.save(filename="../../Downloads/Telegram Desktop/Исходные материалы/Конвертор/export_spreedsheet.xlsx")
